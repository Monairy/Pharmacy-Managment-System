from openpyxl import Workbook,load_workbook
from tkinter import *

Path = ("/home/ahmad/Downloads/ASU FOE/Software Engineering/Project/MedicineDatabase.xlsx") #Path Of DataBase
Path2 = ("/home/ahmad/Downloads/ASU FOE/Software Engineering/Project/systemUsersDatabase.xlsx")

"""
Notes:
   for writing in cell; column=letter, row=number(starting with 1)
   save is a must after writing
"""
  
class MedicineDatabase:
  DataBaseFile=Path
  workbook = load_workbook(filename=DataBaseFile) #Write
  sheet= workbook.active #Write
  
  def GetRowIndex(self): #index of row to write in
       return self.sheet.max_row+1
  def SaveDatabase(self):
      self.workbook.save(filename=self.DataBaseFile)
  def SearchMedicineByName(self,MedicineName): # Find Medicine in Database with it's name and return dictionary of data
    DictOfData={}         
    for row in range(1,self.sheet.max_row+1): # loop over rows
       if(str(self.sheet["A"+str(row)].value).lower()==MedicineName.lower()):
             DictOfData["barcode"]= self.sheet["B"+str(row)].value
             DictOfData["quantity"]= self.sheet["C"+str(row)].value
             DictOfData["expire"]= self.sheet["D"+str(row)].value
             DictOfData["price"]= self.sheet["E"+str(row)].value
             return DictOfData
  def EditMedicineBarcode(self,MedicineName,NewBarcode):
    for row in range(1,self.sheet.max_row+1): # loop over rows
       if(str(self.sheet["A"+str(row)].value).lower()==MedicineName.lower()):
            self.sheet["B"+str(row)]= NewBarcode #update barcode
            self.SaveDatabase()
  def EditMedicineQuantity(self,MedicineName,NewQuantity):
    for row in range(1,self.sheet.max_row+1): # loop over rows
       if(str(self.sheet["A"+str(row)].value).lower()==MedicineName.lower()):
            self.sheet["C"+str(row)]= NewQuantity #update quantity
            self.SaveDatabase()
  def EditMedicineExpire(self,MedicineName,NewExpire):
    for row in range(1,self.sheet.max_row+1): # loop over rows
       if(str(self.sheet["A"+str(row)].value).lower()==MedicineName.lower()):
            self.sheet["D"+str(row)]= NewExpire #update expire
            self.SaveDatabase()            
  def EditMedicinePrice(self,MedicineName,NewPrice):
    for row in range(1,self.sheet.max_row+1): # loop over rows
       if(str(self.sheet["A"+str(row)].value).lower()==MedicineName.lower()):
            self.sheet["E"+str(row)]= NewPrice #update price
            self.SaveDatabase()
       
class systemUsersDatabase:
   DataBaseFile = Path2
   workbook2 = load_workbook(filename = DataBaseFile) #Write
   sheet2 = workbook2.active #Write

   def checkLogin(self,systemUserName, systemUserPass):     
      for row in range(1,self.sheet2.max_row+1): # loop over rows
         if(str(self.sheet2["A"+str(row)].value).lower()==systemUserName.lower() and str(self.sheet2["B"+str(row)].value).lower()==systemUserPass.lower()):
            return TRUE
'''
class systemUser:
   name="null"
   password="null"
   def enterName(self,Name):
      self.name=Name.lower()
   def enterPass(self,Pass):
      self.password=Pass
'''
class Medicine:
 name="null"
 barcode="null"
 quantity="null"
 expire="null"
 price="null"
 def SetName(self,Name):
    self.name=Name.lower()
 def SetBarcode(self,Barcode):
    self.barcode=Barcode
 def SetQuantity(self,Quantity):
    self.quantity=Quantity    
 def SetExpire(self,Expire):
    self.expire=Expire    
 def SetPrice(self,Price):
    self.price=Price
    
 def AddToDataBase(self):
    DB=MedicineDatabase()
    RowIndex=str(DB.GetRowIndex())
    if (self.name!=""): # To Avoid storing empty objects
     DB.sheet["A"+RowIndex]= self.name
     DB.sheet["B"+RowIndex]= self.barcode
     DB.sheet["C"+RowIndex]= self.quantity
     DB.sheet["D"+RowIndex]= self.expire
     DB.sheet["E"+RowIndex]= self.price
     DB.SaveDatabase()
    


def AddMedicineUI():
    DestroyAll()
    
    global label1,label2,label3,label4,label5,entry1,entry2,entry3,entry4,entry5,ButtonAdd
    
    label1= Label(GUI,text="name",bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    label1.place(x=0,y=120)
    label2= Label(GUI,text="barcode",bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    label2.place(x=0,y=160)
    label3= Label(GUI,text="quantity",bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    label3.place(x=0,y=200)
    label4= Label(GUI,text="expire",bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    label4.place(x=0,y=240)
    label5= Label(GUI,text="price",bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    label5.place(x=0,y=280)
    
    entry1=Entry(GUI , font=("Times", 20))
    entry1.place(x=150,y=120)
    entry2=Entry(GUI , font=("Times", 20))
    entry2.place(x=150,y=160)
    entry3=Entry(GUI , font=("Times", 20))
    entry3.place(x=150,y=200)
    entry4=Entry(GUI , font=("Times", 20))
    entry4.place(x=150,y=240)
    entry5=Entry(GUI , font=("Times", 20))
    entry5.place(x=150,y=280)

    ButtonAdd = Button(GUI, text ="Add",font=("Arial", 14),command = lambda : NewMedicine())
    ButtonAdd.configure(height=1,width=10)
    ButtonAdd.place(x=100,y=320)
    
    
def NewMedicine():
  Med=Medicine()
  Med.SetName(entry1.get())
  Med.SetBarcode(entry2.get())
  Med.SetQuantity(entry3.get())   
  Med.SetExpire(entry4.get())   
  Med.SetPrice(entry5.get())
  Med.AddToDataBase()
  labeldone= Label(GUI,text="Medicine Added Successfuly",bg="GREY",fg="RED",font=("Times", 20)) 
  labeldone.place(x=0,y=360)
  GUI.after(2000,lambda:labeldone.destroy()) # done label appear and disappear after time
'''
def NewLogin():
  sysUser=systemUser()
  systemUser.enterName(usernameEntry.get())
  systemUser.enterPass(passwordEntry.get())
  labeldone= Label(GUI,text="Medicine Added Successfuly",bg="GREY",fg="RED",font=("Times", 20)) 
  labeldone.place(x=0,y=360)
  GUI.after(2000,lambda:labeldone.destroy()) # done label appear and disappear after time
'''

def newLogin():
  DB2 = systemUsersDatabase()
  if(DB2.checkLogin(usernameEntry.get(), passwordEntry.get()) == True):
   successfulLabel= Label(LoginScreen,text="Successful Login",bg="GREY",fg="RED",font=("Times", 15))
   successfulLabel.place(x=800,y=400)
   main()
  else:
   failLabel= Label(LoginScreen,text="Failure Login",bg="GREY",fg="RED",font=("Times", 15))
   failLabel.place(x=800,y=400)

def DestroyAll(): # make sure that area we use is clear before placing objects
  try:
        label1.destroy()
        label2.destroy()
        label3.destroy()
        label4.destroy()
        label5.destroy()
        entry1.destroy()
        entry2.destroy()
        entry3.destroy()
        entry4.destroy()
        entry5.destroy()
        ButtonAdd.destroy()
        label6.destroy()
        entry6.destroy()
        ButtonSearch.destroy()
        entry7.destroy()
        entry8.destroy()
        entry9.destroy()
        entry10.destroy()
        ButtonBarcode.destroy()
        ButtonQuantity.destroy()
        ButtonExpire.destroy()
        ButtonPrice.destroy()
  except:
       pass


def EditExistingMedicineUI(): # enter name, click search then show result of search by: showDataUI()
    DestroyAll()       
    global label6,entry6,ButtonSearch
    label6= Label(GUI,text="Enter Medicine Name: ",bg="LightBlue",fg="white",font=("Times", 20),width=20,relief="ridge")
    label6.place(x=200,y=120)
    
    entry6=Entry(GUI , font=("Times", 20),width=20)
    entry6.place(x=210,y=160)

    
    ButtonSearch = Button(GUI, text ="Search",font=("Arial", 14),command = lambda : ShowDataUI())
    ButtonSearch.configure(height=1,width=10)
    ButtonSearch.place(x=290,y=200)


def ShowDataUI(): #After Search
   global entry7,entry8,entry9,entry10,ButtonBarcode,ButtonQuantity,ButtonExpire,ButtonPrice     
   if (entry6.get()!=""):
     DB = MedicineDatabase()
     MedicineData = DB.SearchMedicineByName(entry6.get())  #dictionary of medicine data from database

     entry7 = Entry(GUI, font=("Times", 15),width=15) # entry for barcode
     entry7.insert(0,MedicineData.get("barcode"))
     entry7.place(x=200,y=320)

     ButtonBarcode = Button(GUI, text ="Edit Barcode",font=("Arial", 12),command=lambda:EditMedicine(1))
     ButtonBarcode.configure(height=1,width=15)
     ButtonBarcode.place(x=205,y=360)
####################################
     entry8 = Entry(GUI, font=("Times", 15),width=15) # entry for quantity
     entry8.insert(0,MedicineData.get("quantity"))
     entry8.place(x=400,y=320)

     ButtonQuantity = Button(GUI, text ="Edit Quantity",font=("Arial", 12),command=lambda:EditMedicine(2))
     ButtonQuantity.configure(height=1,width=15)
     ButtonQuantity.place(x=405,y=360)
####################################
     entry9 = Entry(GUI, font=("Times", 15),width=15) # entry for expire
     entry9.insert(0,MedicineData.get("expire"))
     entry9.place(x=600,y=320)

     ButtonExpire = Button(GUI, text ="Edit Expire",font=("Arial", 12),command=lambda:EditMedicine(3))
     ButtonExpire.configure(height=1,width=15)
     ButtonExpire.place(x=605,y=360)
####################################
     entry10 = Entry(GUI, font=("Times", 15),width=15) # entry for price
     entry10.insert(0,MedicineData.get("price"))
     entry10.place(x=800,y=320)

     ButtonPrice = Button(GUI, text ="Edit Price",font=("Arial", 12),command=lambda:EditMedicine(4))
     ButtonPrice.configure(height=1,width=15)
     ButtonPrice.place(x=805,y=360)

     
def EditMedicine(arg): # 1:barcode, 2:quantity, 3:expire, 4:price
    DB = MedicineDatabase()
    MedicineData = DB.SearchMedicineByName(entry6.get())  #search in db and get dictionary of medicine data from database

    if (arg==1):
      DB.EditMedicineBarcode(entry6.get(),entry7.get()) # EditMedicineBarcode(name,barcode)

      label= Label(GUI,text="Barcode Modified Successfuly",bg="GREY",fg="RED",font=("Times", 15))
      label.place(x=200,y=400)
      GUI.after(2000,lambda:label.destroy())     # label appear and disappear after time

    if (arg==2):
      DB.EditMedicineQuantity(entry6.get(),entry8.get()) # EditMedicineQuantity(name,quantity)

      label= Label(GUI,text="Quantity Modified Successfuly",bg="GREY",fg="RED",font=("Times", 15))
      label.place(x=400,y=400)
      GUI.after(2000,lambda:label.destroy())    

    if (arg==3):
      DB.EditMedicineExpire(entry6.get(),entry9.get()) # EditMedicineExpire(name,expire)

      label= Label(GUI,text="Expire Modified Successfuly",bg="GREY",fg="RED",font=("Times", 15))
      label.place(x=600,y=400)
      GUI.after(2000,lambda:label.destroy())

    if (arg==4):
      DB.EditMedicinePrice(entry6.get(),entry10.get()) # EditMedicinePrice(name,price)

      label= Label(GUI,text="Price Modified Successfuly",bg="GREY",fg="RED",font=("Times", 15))
      label.place(x=800,y=400)
      GUI.after(2000,lambda:label.destroy())

def main():
 global GUI
 GUI = Tk()
 GUI.title("Pharmacy Managment System")
 GUI.configure(bg='GREY')
 GUI.minsize(1400,650)
 GUI.resizable(0,0)
 
 labelbanner= Label(GUI,text="Pharmacy Managment System",bg="LightBlue",fg="white",font=("Times", 30))
 labelbanner.grid(columnspan=7,padx=500)
 LoginScreen.destroy()


 B0 = Button(GUI, text ="Add New Medicine", font=("Arial", 15),command = lambda : AddMedicineUI())
 B0.configure(height=2,width=16)
 B0.grid(row=1,column=0)

 B1 = Button(GUI, text ="Edit Existing Medicine",font=("Arial", 15), command =lambda :  EditExistingMedicineUI())
 B1.configure(height=2,width=17)
 B1.grid(row=1,column=1)

 B2 = Button(GUI, text ="Ay 7aga bardo",font=("Arial", 15), command =lambda :  AddMedicineUI())
 B2.configure(height=2,width=16)
 B2.grid(row=1,column=2)

 B3 = Button(GUI, text ="Ay 7aga bardo",font=("Arial", 15), command =lambda :  AddMedicineUI())
 B3.configure(height=2,width=16)
 B3.grid(row=1,column=3)

 B4 = Button(GUI, text ="Ay 7aga bardo",font=("Arial", 15), command =lambda :  AddMedicineUI())
 B4.configure(height=2,width=16)
 B4.grid(row=1,column=4)

 B5 = Button(GUI, text ="Ay 7aga bardo",font=("Arial", 15), command =lambda :  AddMedicineUI())
 B5.configure(height=2,width=16)
 B5.grid(row=1,column=5)

 B6 = Button(GUI, text ="Ay 7aga bardo",font=("Arial", 15), command =lambda :  AddMedicineUI())
 B6.configure(height=2,width=16)
 B6.grid(row=1,column=6)

 labelfooter= Label(GUI,text="Version 1.00",bg="grey",font=("Times", 14))
 labelfooter.place(x=700,y=600)

 
 GUI.mainloop()

 

def login():
  global LoginScreen, usernameEntry, passwordEntry
  LoginScreen = Tk()
  LoginScreen.title("Pharmacy Managment System")
  LoginScreen.configure(bg='GREY')
  LoginScreen.minsize(1400,650)
  LoginScreen.resizable(0,0)
  
  labelbanner= Label(LoginScreen,text="Login To Continue",bg="RED",fg="BLACK",font=("Times", 30))
  labelbanner.grid(columnspan=7,padx=500)

  usernameLabel=Label(LoginScreen, text=" Username: ", width=7, bg="grey",font=("Times",20))
  usernameLabel.place(x=350,y=150)
  usernameEntry = Entry(LoginScreen,font=("Times", 20))
  usernameEntry.place(x=550,y=150)
  
  PasswordLabel=Label(LoginScreen, text="Password:" , width=7, bg="grey",font=("Times",20))
  PasswordLabel.place(x=350,y=250)
  passwordEntry = Entry(LoginScreen, font=("Times", 20), show= '*')
  passwordEntry.place(x=550,y=250)

  loginbutton= Button(LoginScreen, text ="Login",font=("Arial", 20), command =lambda : newLogin())
  loginbutton.place(x=600,y=400)
  LoginScreen.mainloop()
login()


