from openpyxl import Workbook, load_workbook
from tkinter import *


Path = ("MedicineDatabase.xlsx")  # Path Of DataBase

Admin_path = ('EmployeeDatabase.xlsx')

"""
Notes:
   for writing in cell; column=letter, row=number(starting with 1)
   save is a must after writing
"""


class Database:
    DataBaseFile = Path
    workbook = load_workbook(filename=DataBaseFile)  # Write
    sheet = workbook.active  # Write

    def GetRowIndex(self):  # index of row to write in
        return self.sheet.max_row + 1

    def SaveDatabase(self):
        self.workbook.save(filename=self.DataBaseFile)


class EmployeeDataBase:
    DataBase = Admin_path
    wb = load_workbook(filename=DataBase)
    sheet1 = wb.active

    def GetRow(self):
        return self.sheet1.max_row + 1

    def SaveEmployeeDatabase(self):
        self.wb.save(filename=self.DataBase)

    def SearchEmployee(self, EmployeeName):
        EmployeeDectionary = {}
        for row in range(1, self.sheet1.max_row + 1):  # loop over rows
            if (str(self.sheet1["C" + str(row)].value).lower() == EmployeeName.lower()):
                EmployeeDectionary["username"] = self.sheet1["A" + str(row)].value
                EmployeeDectionary["password"] = self.sheet1["B" + str(row)].value
                EmployeeDectionary["national_id"] = self.sheet1["D" + str(row)].value
                EmployeeDectionary["address"] = self.sheet1["E" + str(row)].value
                EmployeeDectionary["phone"] = self.sheet1["F" + str(row)].value
                EmployeeDectionary["age"] = self.sheet1["G" + str(row)].value
                return EmployeeDectionary


class Medicine:
    name = "null"
    barcode = "null"
    quantity = "null"
    expire = "null"
    price = "null"

    def SetName(self, Name):
        self.name = Name.lower()

    def SetBarcode(self, Barcode):
        self.barcode = Barcode

    def SetQuantity(self, Quantity):
        self.quantity = Quantity

    def SetExpire(self, Expire):
        self.expire = Expire

    def SetPrice(self, Price):
        self.price = Price

    def AddToDataBase(self):
        DB = Database()
        RowIndex = str(DB.GetRowIndex())
        DB.sheet["A" + RowIndex] = self.name
        DB.sheet["B" + RowIndex] = self.barcode
        DB.sheet["C" + RowIndex] = self.quantity
        DB.sheet["D" + RowIndex] = self.expire
        DB.sheet["E" + RowIndex] = self.price
        DB.SaveDatabase()


class Employee:
    name = 'null'
    national_id = 'null'
    phone_number = 'null'
    address = 'null'
    age = 'null'
    username = 'null'
    password = 'null'

    def SetUserName(self, UserName):
        self.username = UserName.lower()

    def SetPassword(self, Password):
        self.password = Password.lower()

    def SetName(self, Name):
        self.name = Name.lower()

    def SetNationalId(self, Id):
        self.national_id = Id

    def SetPhoneNumber(self, phone):
        self.phone_number = phone

    def SetAddress(self, Add):
        self.address = Add

    def SetAge(self, Age):
        self.age = Age

    def GetName(self):
        return self.name

    def GetNationalId(self):
        return self.national_id

    def GetPhoneNumber(self):
        return self.phone_number

    def GetAddress(self):
        return self.address

    def GetAge(self):
        return self.age

    def AddDataBase(self):
        AdminDB = EmployeeDataBase()
        Row = str(AdminDB.GetRow())
        AdminDB.sheet1["A" + Row] = self.username
        AdminDB.sheet1["B" + Row] = self.password
        AdminDB.sheet1["C" + Row] = self.name
        AdminDB.sheet1["D" + Row] = self.national_id
        AdminDB.sheet1["E" + Row] = self.phone_number
        AdminDB.sheet1["F" + Row] = self.address
        AdminDB.sheet1["G" + Row] = self.age
        AdminDB.SaveEmployeeDatabase()





def NewEmployee():
    employee = Employee()
    employee.SetUserName(entry6.get())
    employee.SetPassword(entry7.get())
    employee.SetName(entry8.get())
    employee.SetNationalId(entry9.get())
    employee.SetAddress(entry8.get())
    employee.SetPhoneNumber(entry10.get())
    employee.SetAge(entry11.get())
    employee.AddDataBase()
    HideEmployeeUI()


def HideEmployeeUI():
    global employeeadded
    label6.destroy()
    label7.destroy()
    label8.destroy()
    label9.destroy()
    labe20.destroy()
    labe21.destroy()
    labe22.destroy()
    entry6.destroy()
    entry7.destroy()
    entry8.destroy()
    entry9.destroy()
    entry10.destroy()
    entry11.destroy()
    entry12.destroy()
    AddInfo.destroy()
    employeeadded = Label(AdminGui, text="Employee Added Successfuly", bg="GREY", fg="RED", font=("Times", 20))
    employeeadded.place(x=120, y=200)



def DisplayEmployeeInfo(searched_name):
    global info_display
    try:
        label6.destroy()
        label7.destroy()
        label8.destroy()
        label9.destroy()
        labe20.destroy()
        labe21.destroy()
        labe22.destroy()
        entry6.destroy()
        entry7.destroy()
        entry8.destroy()
        entry9.destroy()
        entry10.destroy()
        entry11.destroy()
        entry12.destroy()
        AddInfo.destroy()
        employeeadded.destroy()
    except:
        pass
    D2 = EmployeeDataBase()
    Dictdata = D2.SearchEmployee(searched_name)
    info_display = Text(AdminGui, height=15, width=25)
    info_display.grid(column=2, row=5)
    info_display.insert(INSERT, 'Name: '+searched_name+'\n')
    info_display.insert(INSERT, 'User Name: ' + Dictdata["username"] + '\n')
    info_display.insert(INSERT, 'Password: ' + Dictdata["password"] + '\n')
    info_display.insert(INSERT, 'National ID: ' + Dictdata["national_id"] + '\n')
    info_display.insert(INSERT, 'Address: ' + Dictdata["address"] + '\n')
    info_display.insert(INSERT, 'Phone: ' + Dictdata["phone"] + '\n')
    info_display.insert(INSERT, 'Age: ' + Dictdata["age"] + '\n')
    info_display.configure(state='disabled')



def GetEmployeeName():
    global search_label,search_entry,Buttonseach
    try:
        label6.destroy()
        label7.destroy()
        label8.destroy()
        label9.destroy()
        labe20.destroy()
        labe21.destroy()
        labe22.destroy()
        entry6.destroy()
        entry7.destroy()
        entry8.destroy()
        entry9.destroy()
        entry10.destroy()
        entry11.destroy()
        entry12.destroy()
        AddInfo.destroy()
        employeeadded.destroy()
    except:
        pass
    search_label = Label(AdminGui, text="Name", bg="LightBlue", fg="white", font=("Times", 20), width=9, relief="ridge")
    search_label.place(x=0, y=120)
    search_entry = Entry(AdminGui, font=("Times", 20))
    search_entry.place(x=150, y=120)
    Buttonseach = Button(AdminGui, text="Search", font=("Arial", 14), command=lambda: DisplayEmployeeInfo(search_entry.get()))
    Buttonseach.configure(height=1, width=10)
    Buttonseach.place(x=100, y=160)

def AddEmployee():
    global label6, label7, label8, label9, labe20,labe21,labe22, entry6, entry7, entry8, entry9, entry10,entry11,entry12, AddInfo
    try:
        Buttonseach.destroy()
        search_entry.destroy()
        search_label.destroy()
        info_display.destroy()
    except:
        pass
    label6 = Label(AdminGui, text="User Name", bg="LightBlue", fg="white", font=("Times", 20), width=9, relief="ridge")
    label6.place(x=0, y=120)
    label7 = Label(AdminGui, text="Password", bg="LightBlue", fg="white", font=("Times", 20), width=9, relief="ridge")
    label7.place(x=0, y=160)
    label8 = Label(AdminGui, text="Name", bg="LightBlue", fg="white", font=("Times", 20), width=9, relief="ridge")
    label8.place(x=0, y=200)
    label9 = Label(AdminGui, text="National ID", bg="LightBlue", fg="white", font=("Times", 20), width=9, relief="ridge")
    label9.place(x=0, y=240)
    labe20 = Label(AdminGui, text="Address", bg="LightBlue", fg="white", font=("Times", 20), width=9, relief="ridge")
    labe20.place(x=0, y=280)
    labe21 = Label(AdminGui, text="Phone Num", bg="LightBlue", fg="white", font=("Times", 20), width=9, relief="ridge")
    labe21.place(x=0, y=320)
    labe22 = Label(AdminGui, text="Age", bg="LightBlue", fg="white", font=("Times", 20), width=9, relief="ridge")
    labe22.place(x=0, y=360)



    entry6 = Entry(AdminGui, font=("Times", 20))
    entry6.place(x=150, y=120)
    entry7 = Entry(AdminGui, font=("Times", 20))
    entry7.place(x=150, y=160)
    entry8 = Entry(AdminGui, font=("Times", 20))
    entry8.place(x=150, y=200)
    entry9 = Entry(AdminGui, font=("Times", 20))
    entry9.place(x=150, y=240)
    entry10 = Entry(AdminGui, font=("Times", 20))
    entry10.place(x=150, y=280)
    entry11 = Entry(AdminGui, font=("Times", 20))
    entry11.place(x=150, y=320)
    entry12 = Entry(AdminGui, font=("Times", 20))
    entry12.place(x=150, y=360)

    AddInfo = Button(AdminGui, text="Add", font=("Arial", 14), command=lambda: NewEmployee())
    AddInfo.configure(height=1, width=10)
    AddInfo.place(x=100, y=400)




def DestroyAdminUi():
    AdminGui.destroy()

def AdminUi():
    global AdminGui
    AdminGui = Tk()


    AdminGui.title("Adminstration mode")
    AdminGui.configure(bg='GREY')
    AdminGui.minsize(1400, 650)
    AdminGui.resizable(0, 0)
    banner = Label(AdminGui, text="Adminstration mode", bg="LightBlue", fg="white", font=("Times", 30))
    banner.grid(columnspan=7, padx=500)

    global employeeadded
    employeeadded = Label(AdminGui, text="Employee Added Successfuly", bg="GREY", fg="RED", font=("Times", 20))

    Bot1 = Button(AdminGui, text="Add New Employee", font=("Arial", 10), command=lambda: AddEmployee())
    Bot1.configure(height=3, width=20)
    Bot1.grid(row=1, column=0)

    Bot2 = Button(AdminGui, text="Get Employee Information", font=("Arial", 10), command=lambda: GetEmployeeName())
    Bot2.configure(height=3, width=20)
    Bot2.grid(row=1, column=1)

    Bot3 = Button(AdminGui, text="Back", font=("Arial", 10), command=lambda:  DestroyAdminUi())
    Bot3.configure(height=3, width=20)
    Bot3.grid(row=1, column=2)

    AdminGui.mainloop()



'''def AdminLogin():
    global adminusernameEntry, adminpasswordEntry
    global AdminLoginGUI
    AdminLoginGUI = Tk()
    AdminLoginGUI.title("Admin login")
    AdminLoginGUI.configure(bg='GREY')
    AdminLoginGUI.minsize(1400, 650)
    AdminLoginGUI.resizable(0, 0)
    adminbanner = Label(AdminLoginGUI, text="Admin login", bg="LightBlue", fg="white", font=("Times", 30))
    AdminLoginGUI.grid(columnspan=7, padx=500)

    adminusernameLabel = Label(AdminLoginGUI, text=" Username: ", width=7, bg="grey", font=("Times", 20))
    adminusernameLabel.place(x=350, y=150)
    adminusernameEntry = Entry(AdminLoginGUI, font=("Times", 20))
    adminusernameEntry.place(x=550, y=150)

    adminPasswordLabel = Label(AdminLoginGUI, text="Password:", width=7, bg="grey", font=("Times", 20))
    adminPasswordLabel.place(x=350, y=250)
    adminpasswordEntry = Entry(AdminLoginGUI, font=("Times", 20), show='*')
    adminpasswordEntry.place(x=550, y=250)

    adminloginbutton = Button(AdminLoginGUI, text="Login", font=("Arial", 20), command=lambda: check())
    adminloginbutton.place(x=600, y=400)
    AdminLoginGUI.mainloop()'''


'''def check():
    Admindata = EmployeeDataBase()
    if (Admindata.sheet1['A1'] == adminusernameEntry.get()) and (Admindata.sheet1['B1'] == adminpasswordEntry.get()):
        AdminUi()

    else:
        labelerror = Label(AdminLoginGUI, text="User name or password is wrong", bg="GREY", fg="RED", font=("Times", 20))
        labelerror.place(x=150, y=200)
        AdminLogin()'''





def AddMedicineUI():
    global label1, label2, label3, label4, label5, entry1, entry2, entry3, entry4, entry5, ButtonAdd
    labeldone.destroy()
    label1 = Label(GUI, text="Name", bg="LightBlue", fg="white", font=("Times", 20), width=7, relief="ridge")
    label1.place(x=0, y=120)
    label2 = Label(GUI, text="Barcode", bg="LightBlue", fg="white", font=("Times", 20), width=7, relief="ridge")
    label2.place(x=0, y=160)
    label3 = Label(GUI, text="Quantity", bg="LightBlue", fg="white", font=("Times", 20), width=7, relief="ridge")
    label3.place(x=0, y=200)
    label4 = Label(GUI, text="Expire", bg="LightBlue", fg="white", font=("Times", 20), width=7, relief="ridge")
    label4.place(x=0, y=240)
    label5 = Label(GUI, text="Price", bg="LightBlue", fg="white", font=("Times", 20), width=7, relief="ridge")
    label5.place(x=0, y=280)

    entry1 = Entry(GUI, font=("Times", 20))
    entry1.place(x=150, y=120)
    entry2 = Entry(GUI, font=("Times", 20))
    entry2.place(x=150, y=160)
    entry3 = Entry(GUI, font=("Times", 20))
    entry3.place(x=150, y=200)
    entry4 = Entry(GUI, font=("Times", 20))
    entry4.place(x=150, y=240)
    entry5 = Entry(GUI, font=("Times", 20))
    entry5.place(x=150, y=280)

    ButtonAdd = Button(GUI, text="Add", font=("Arial", 14), command=lambda: NewMedicine())
    ButtonAdd.configure(height=1, width=10)
    ButtonAdd.place(x=100, y=320)


def NewMedicine():
    Med = Medicine()
    Med.SetName(entry1.get())
    Med.SetBarcode(entry2.get())
    Med.SetQuantity(entry3.get())
    Med.SetExpire(entry4.get())
    Med.SetPrice(entry5.get())
    Med.AddToDataBase()
    HideMedicineUI()


def HideMedicineUI():
    global labeldone
    label1.destroy()
    label2.destroy()
    label3.destroy()
    label4.destroy()
    label5.destroy()
    entry1.destroy()
    entry2.destroy()
    entry3.destroy()
    entry4.destroy()
    entry5.destroy()
    ButtonAdd.destroy()
    labeldone = Label(GUI, text="Medicine Added Successfuly", bg="GREY", fg="RED", font=("Times", 20))
    labeldone.place(x=120, y=200)




def main():
    global GUI
    LoginScreen.destroy()
    GUI = Tk()
    GUI.title("Pharmacy Managment System")
    GUI.configure(bg='GREY')
    GUI.minsize(1400, 650)
    GUI.resizable(0, 0)
    labelbanner = Label(GUI, text="Pharmacy Managment System", bg="LightBlue", fg="white", font=("Times", 30))
    labelbanner.grid(columnspan=7, padx=500)

    global labeldone
    labeldone = Label(GUI, text="Medicine Added Successfuly", bg="GREY", fg="RED", font=("Times", 20))

    B0 = Button(GUI, text="Add New Medicine", font=("Arial", 10), command=lambda: AddMedicineUI())
    B0.configure(height=3, width=20)
    B0.grid(row=1, column=0)

    B1 = Button(GUI, text="Ay 7aga", font=("Arial", 10), command=lambda: AddMedicineUI())
    B1.configure(height=3, width=20)
    B1.grid(row=1, column=1)

    B2 = Button(GUI, text="Adminstration", font=("Arial", 10), command=lambda: AdminUi())
    B2.configure(height=3, width=20)
    B2.grid(row=1, column=2)

    B3 = Button(GUI, text="Ay 7aga bardo", font=("Arial", 10), command=lambda: AddMedicineUI())
    B3.configure(height=3, width=20)
    B3.grid(row=1, column=3)

    B4 = Button(GUI, text="Ay 7aga bardo", font=("Arial", 10), command=lambda: AddMedicineUI())
    B4.configure(height=3, width=20)
    B4.grid(row=1, column=4)

    B5 = Button(GUI, text="Ay 7aga bardo", font=("Arial", 10), command=lambda: AddMedicineUI())
    B5.configure(height=3, width=20)
    B5.grid(row=1, column=5)

    B6 = Button(GUI, text="Ay 7aga bardo", font=("Arial", 10), command=lambda: AddMedicineUI())
    B6.configure(height=3, width=20)
    B6.grid(row=1, column=6)

    labelfooter = Label(GUI, text="Version 1.00", bg="grey", font=("Times", 14))
    labelfooter.place(x=700, y=600)
    GUI.mainloop()


def login():
    global LoginScreen
    LoginScreen = Tk()
    LoginScreen.title("Pharmacy Managment System")
    LoginScreen.configure(bg='GREY')
    LoginScreen.minsize(1400, 650)
    LoginScreen.resizable(0, 0)

    labelbanner = Label(LoginScreen, text="Login To Continue", bg="RED", fg="BLACK", font=("Times", 30))
    labelbanner.grid(columnspan=7, padx=500)

    usernameLabel = Label(LoginScreen, text=" Username: ", width=7, bg="grey", font=("Times", 20))
    usernameLabel.place(x=350, y=150)
    usernameEntry = Entry(LoginScreen, font=("Times", 20))
    usernameEntry.place(x=550, y=150)

    PasswordLabel = Label(LoginScreen, text="Password:", width=7, bg="grey", font=("Times", 20))
    PasswordLabel.place(x=350, y=250)
    passwordEntry = Entry(LoginScreen, font=("Times", 20), show='*')
    passwordEntry.place(x=550, y=250)

    loginbutton = Button(LoginScreen, text="Login", font=("Arial", 20), command=lambda: main())
    loginbutton.place(x=600, y=400)
    LoginScreen.mainloop()


login()

"""Medicines = []
i=0
while True:
    print "Enter Medicine Name:  "
    MedName = raw_input()

    if (MedName == 'done'):
        break

    Medicines.append(MedName)
    Medicines[i] = Medicine()
    Medicines[i].SetName(MedName)
    Medicines[i].AddToDataBase()
    i=i+1


    print "Medicine %s is added to Database" % MedName

EditPrice("ketofan","76")
print "Price Updated"
"""
