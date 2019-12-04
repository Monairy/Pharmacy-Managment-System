from openpyxl import Workbook,load_workbook
from tkinter import *
from tkinter import messagebox
import os
import datetime
from tkdocviewer import DocViewer

Path = ("MedicineDatabase.xlsx") #Path Of DataBase
Path2 = ("OrderDatabase.xlsx")
path3=("ClientDatabase.xlsx")
Admin_path = ('EmployeeDatabase.xlsx')

"""
Notes:
   for writing in cell; column=letter, row=number(starting with 1)
   save is a must after writing
"""
class Database:
  def GetRowIndex(self): #index of row to write in
       return self.sheet.max_row+1
  def SaveDatabase(self):
      self.workbook.save(filename=self.DataBaseFile)
      
####################################
########_Medicine_Database##########
####################################
class MedicineDatabase(Database):
  DataBaseFile=Path
  workbook = load_workbook(filename=DataBaseFile) #Write
  sheet= workbook.active #Write
  
  def SearchMedicineByName(self,MedicineName): # Find Medicine in Database with it's name and return dictionary of data
      DictOfData={}
      found = 0
      for row in range(1,self.sheet.max_row+1):# loop over rows
             if(str(self.sheet["A"+str(row)].value).lower()==MedicineName.lower()):
                 DictOfData["barcode"]= self.sheet["B"+str(row)].value
                 DictOfData["quantity"]= self.sheet["C"+str(row)].value
                 DictOfData["expire"]= self.sheet["D"+str(row)].value
                 DictOfData["price"]= self.sheet["E"+str(row)].value
                 found = 1
      if (found == 1):
        return DictOfData
      else:
         DictOfData["barcode"] = "Not Found"
         DictOfData["quantity"]= "Not Found"
         DictOfData["expire"]  = "Not Found"
         DictOfData["price"]   = "Not Found"
         ShowError("Medicine Not Found")
         return DictOfData

  def EditMedicineBarcode(self,MedicineName,NewBarcode):
    for row in range(1,self.sheet.max_row+1): # loop over rows
       if(str(self.sheet["A"+str(row)].value).lower()==MedicineName.lower()):
            self.sheet["B"+str(row)]= NewBarcode #update barcode
            self.SaveDatabase()
  def EditMedicineQuantity(self,MedicineName,NewQuantity):
    for row in range(1,self.sheet.max_row+1): # loop over rows
       if(str(self.sheet["A"+str(row)].value).lower()==MedicineName.lower()):
            self.sheet["C"+str(row)]= NewQuantity #update quantity
            self.SaveDatabase()
  def EditMedicineExpire(self,MedicineName,NewExpire):
    for row in range(1,self.sheet.max_row+1): # loop over rows
       if(str(self.sheet["A"+str(row)].value).lower()==MedicineName.lower()):
            self.sheet["D"+str(row)]= NewExpire #update expire
            self.SaveDatabase()            
  def EditMedicinePrice(self,MedicineName,NewPrice):
    for row in range(1,self.sheet.max_row+1): # loop over rows
       if(str(self.sheet["A"+str(row)].value).lower()==MedicineName.lower()):
            self.sheet["E"+str(row)]= NewPrice #update price
            self.SaveDatabase()
  def AddMedicineToDataBase(self,medicine):
    RowIndex=str(self.GetRowIndex())
    if (medicine.name!=""): # To Avoid storing empty objects
       for row in range(1,self.sheet.max_row):
          if (str(self.sheet["A"+str(row)].value)==medicine.name):
            ShowError("Medicine "+ medicine.name+" is already stored. Use edit mode instead.")
            return 0
       self.sheet["A"+RowIndex]= medicine.name
       self.sheet["B"+RowIndex]= medicine.barcode
       self.sheet["C"+RowIndex]= medicine.quantity
       self.sheet["D"+RowIndex]= medicine.expire
       self.sheet["E"+RowIndex]= medicine.price
       self.SaveDatabase()
       return 1
     
####################################
#########_Order_Database_###########
#################################### 
class OrderDatabase(Database):
  DataBaseFile=Path2
  workbook = load_workbook(filename=DataBaseFile) #Write
  sheet= workbook.active #Write

  def NextReceiptNo(self):
    return self.sheet.max_row
  
  def AddOrderToDataBase(self,order):
    RowIndex=str(self.GetRowIndex())
    self.sheet["A"+RowIndex]= order.receiptnum # OrderID
    self.sheet["B"+RowIndex]="" #initial value to avoid none
    for i in range(0,len(order.items)): #Products 
      self.sheet["B"+RowIndex]=str(self.sheet["B"+RowIndex].value) + order.items[i] + ","
    self.sheet["C"+RowIndex]="" #initial value to avoid none  
    for i in range(0,len(order.items)): #quantities 
      self.sheet["C"+RowIndex]=str(self.sheet["C"+RowIndex].value) + order.quantities[i] + ","
    self.sheet["D"+RowIndex]= order.CalcSum() #total price 
    self.sheet["E"+RowIndex]= order.PaymentType #Payment Type Cash or Visa
    self.sheet["F"+RowIndex]= order.date #Date
    self.sheet["G"+RowIndex]= order.OrderType #order type
    self.sheet["H"+RowIndex]= order.ClientID #Client id
    self.sheet["I"+RowIndex]= order.DeliveryAddress #delivery adress
    
    self.SaveDatabase()
    self.DeductQuantityFromMedicineDataBase(order)

  def DeductQuantityFromMedicineDataBase(self,order): #after successful order
     DB=MedicineDatabase()
     for i in range(0, len(order.items)):
       MedicineName=order.items[i]
       SoldQuantity=int(order.quantities[i])
       AvailableQuantity=int(DB.SearchMedicineByName(MedicineName)["quantity"])
       DB.EditMedicineQuantity(MedicineName,AvailableQuantity-SoldQuantity)

  def DailyProfit(self):
    RowIndex = self.GetRowIndex()
    today = datetime.date.today()
    today =str(today).split()
    DailyProfit=0
    for i in range(1, RowIndex):
        OrderDate =str(self.sheet["F"+str(i)].value).split()

        if OrderDate[0] == today[0]:
            DailyProfit = DailyProfit+int(self.sheet["D"+str(i)].value)
    return DailyProfit
  def MonthlyProfit(self):
    RowIndex = self.GetRowIndex()
    today = datetime.date.today()
    ourmonth = str(today)[0:7]
    MonthlyProfit = 0
    for i in range(1, RowIndex):
        OrderDate = str(self.sheet["F" + str(i)].value)[0:7]

        if (OrderDate == ourmonth):
              MonthlyProfit = MonthlyProfit + int(self.sheet["D" + str(i)].value)
    return MonthlyProfit
  
  def ReturnOrder(self,OrderID):
    for row in range (1,self.sheet.max_row):
      if (str(self.sheet["A"+str(row)].value)==str(OrderID)):
        self.sheet["B"+str(row)]="RETURNED"
        self.sheet["C"+str(row)]="RETURNED"
        self.sheet["D"+str(row)]="0"
        self.SaveDatabase()
  
####################################
##########_Clients_DB_############
####################################
class ClientDatabase(Database):
  DataBaseFile=path3
  workbook = load_workbook(filename=DataBaseFile) #Write
  sheet= workbook.active #Write

  def NextClientID(self):
    return self.sheet.max_row

  def AddClientToDataBase(self,client):
    RowIndex=str(self.GetRowIndex())
    if (client.name!=""): # To Avoid storing empty objects
     self.sheet["A"+RowIndex]= client.ID
     self.sheet["B"+RowIndex]= client.name
     self.sheet["C"+RowIndex]= client.phonenumber
     self.sheet["D"+RowIndex]= client.address
     self.SaveDatabase()

  def GetClientAddress(self,ClientID):
    for i in range (1,self.GetRowIndex()):
      if (str(self.sheet["A"+str(i)].value)==str(ClientID)):
          return self.sheet["D"+str(i)].value  ## returns address of client


####################################
#########_Medicine_Class_###########
####################################
class Medicine:
 name="null"
 barcode="null"
 quantity="null"
 expire="null"
 price="null"
 def SetName(self,Name):
    self.name=Name.lower()
 def SetBarcode(self,Barcode):
    self.barcode=Barcode
 def SetQuantity(self,Quantity):
    self.quantity=Quantity    
 def SetExpire(self,Expire):
    self.expire=Expire    
 def SetPrice(self,Price):
    self.price=Price    
 def AddToDataBase(self):
    DB=MedicineDatabase()
    return DB.AddMedicineToDataBase(self)
    
  
####################################
#########_Receipt_Class_############
####################################   
class Receipt(): # 3 lists items:quantities:prices
  receiptnum="0"
  OrderType="In Store" #store or delivery
  PaymentType="Cash"
  DeliveryAddress="N/A"
  ClientID="0"
  date=str(datetime.datetime.now())[:16]
  items=[]
  quantities=[]
  prices=[]
  def AddItem(self,item,quantity,price):
    self.items.append(item)
    self.quantities.append(quantity)
    self.SetReceiptNum()
    self.prices.append(price)
  def SetReceiptNum(self):
     DB=OrderDatabase()
     self.receiptnum=DB.NextReceiptNo()
  def SetPaymentType(self,paymenttype):
    self.PaymentType=paymenttype    
  def SetType(self,Type):
     self.OrderType=Type   
  def SetDeliveryAddress(self,Address):
     self.DeliveryAddress=Address
  def SetClientID(self,clientid):
     self.ClientID=clientid
     
  def CalcSum(self): # returns string of total order price
    total=0
    for i in range (0,len(self.prices)):
      total = total+int(self.quantities[i])*int(self.prices[i])
    return str(total)
  def AddToDataBase(self):
     DB=OrderDatabase()
     DB.AddOrderToDataBase(self)


####################################
##########_Client_Class_############
####################################
class Client(): 
    ID="null"
    name="null"
    phonenumber="null"
    address="null"

    def SetID(self):
       DB=ClientDatabase()
       self.ID=DB.NextClientID()
    def SetName(self,Name):
       self.name=Name.lower()
    def SetPhone(self,Number):
       self.phonenumber=Number
    def SetAddress(self,Address):
       self.address=Address        
    def AddToDataBase(self):
       DB=ClientDatabase()
       self.SetID()
       DB.AddClientToDataBase(self)
    

####################################
########_USEFUL_FUNCTIONS_##########
#################################### 
    
def ShowError(error):
  errorbox = Tk()
  errorbox.withdraw()
  messagebox.showinfo("Error", error)
  
def space(word,numofspaces):
   space = ""
   for i in range (0,numofspaces-len(word)):
     space = space + " "
   return space
  
####################################
#######_Add_New_Medicine_###########
####################################
def AddMedicineUI():
    global label1,label2,label3,label4,label5,entry1,entry2,entry3,entry4,entry5,ButtonAdd
    DestroyAll()

    label1= Label(GUI,text="name",    bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    label1.place(x=0,y=120)
    label2= Label(GUI,text="barcode", bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    label2.place(x=0,y=160)
    label3= Label(GUI,text="quantity",bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    label3.place(x=0,y=200)
    label4= Label(GUI,text="expire",  bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    label4.place(x=0,y=240)
    label5= Label(GUI,text="price",   bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    label5.place(x=0,y=280)
    
    entry1=Entry(GUI , font=("Times", 20))
    entry1.place(x=150,y=120)
    entry2=Entry(GUI , font=("Times", 20))
    entry2.place(x=150,y=160)
    entry3=Entry(GUI , font=("Times", 20))
    entry3.place(x=150,y=200)
    entry4=Entry(GUI , font=("Times", 20))
    entry4.place(x=150,y=240)
    entry5=Entry(GUI , font=("Times", 20))
    entry5.place(x=150,y=280)

    ButtonAdd = Button(GUI, text ="Add",font=("Arial", 14),command = lambda : NewMedicine())
    ButtonAdd.configure(height=1,width=10)
    ButtonAdd.place(x=100,y=320)
    
    
def NewMedicine():
  Med=Medicine()
  Med.SetName(entry1.get())
  Med.SetBarcode(entry2.get())
  Med.SetQuantity(entry3.get())   
  Med.SetExpire(entry4.get())   
  Med.SetPrice(entry5.get())
  Flag=Med.AddToDataBase()
  if (Flag!=0):
      labeldone= Label(GUI,text="Medicine Added Successfuly",bg="GREY",fg="RED",font=("Times", 20)) 
      labeldone.place(x=0,y=360)
      GUI.after(2000,lambda:labeldone.destroy()) # done label appear and disappear after time

####################################
########_Edit_Medicine_#############
####################################  
def EditExistingMedicineUI(): # enter name, click search then show result of search by: showDataUI()
    DestroyAll()       
    global label6,entry6,ButtonSearch
    label6= Label(GUI,text="Enter Medicine Name: ",bg="LightBlue",fg="white",font=("Times", 20),width=20,relief="ridge")
    label6.place(x=200,y=120)
    
    entry6=Entry(GUI , font=("Times", 20),width=20)
    entry6.place(x=210,y=160)

    
    ButtonSearch = Button(GUI, text ="Search",font=("Arial", 14),command = lambda : ShowDataUI())
    ButtonSearch.configure(height=1,width=10)
    ButtonSearch.place(x=290,y=200)


def ShowDataUI(): #After Search
   global entry7,entry8,entry9,entry10,ButtonBarcode,ButtonQuantity,ButtonExpire,ButtonPrice     
   if (entry6.get()!=""):
     DB = MedicineDatabase()
     MedicineData = DB.SearchMedicineByName(entry6.get())  #dictionary of medicine data from database

     entry7 = Entry(GUI, font=("Times", 15),width=15) # entry for barcode
     entry7.insert(0,MedicineData.get("barcode"))
     entry7.place(x=200,y=320)

     ButtonBarcode = Button(GUI, text ="Edit Barcode",font=("Arial", 12),command=lambda:EditMedicine(1))
     ButtonBarcode.configure(height=1,width=15)
     ButtonBarcode.place(x=205,y=360)
####################################
     entry8 = Entry(GUI, font=("Times", 15),width=15) # entry for quantity
     entry8.insert(0,MedicineData.get("quantity"))
     entry8.place(x=400,y=320)

     ButtonQuantity = Button(GUI, text ="Edit Quantity",font=("Arial", 12),command=lambda:EditMedicine(2))
     ButtonQuantity.configure(height=1,width=15)
     ButtonQuantity.place(x=405,y=360)
####################################
     entry9 = Entry(GUI, font=("Times", 15),width=15) # entry for expire
     entry9.insert(0,MedicineData.get("expire"))
     entry9.place(x=600,y=320)

     ButtonExpire = Button(GUI, text ="Edit Expire",font=("Arial", 12),command=lambda:EditMedicine(3))
     ButtonExpire.configure(height=1,width=15)
     ButtonExpire.place(x=605,y=360)
####################################
     entry10 = Entry(GUI, font=("Times", 15),width=15) # entry for price
     entry10.insert(0,MedicineData.get("price"))
     entry10.place(x=800,y=320)

     ButtonPrice = Button(GUI, text ="Edit Price",font=("Arial", 12),command=lambda:EditMedicine(4))
     ButtonPrice.configure(height=1,width=15)
     ButtonPrice.place(x=805,y=360)

     
def EditMedicine(arg): # execute action depending on arg, 1:barcode, 2:quantity, 3:expire, 4:price
    DB = MedicineDatabase()
    MedicineData = DB.SearchMedicineByName(entry6.get())  #search in db and get dictionary of medicine data from database

    if (arg==1):
      DB.EditMedicineBarcode(entry6.get(),entry7.get()) # EditMedicineBarcode(name,barcode)

      label= Label(GUI,text="Barcode Modified Successfuly",bg="GREY",fg="RED",font=("Times", 15))
      label.place(x=200,y=400)
      GUI.after(2000,lambda:label.destroy())     # label appear and disappear after time

    if (arg==2):
      DB.EditMedicineQuantity(entry6.get(),entry8.get()) # EditMedicineQuantity(name,quantity)

      label= Label(GUI,text="Quantity Modified Successfuly",bg="GREY",fg="RED",font=("Times", 15))
      label.place(x=400,y=400)
      GUI.after(2000,lambda:label.destroy())    

    if (arg==3):
      DB.EditMedicineExpire(entry6.get(),entry9.get()) # EditMedicineExpire(name,expire)

      label= Label(GUI,text="Expire Modified Successfuly",bg="GREY",fg="RED",font=("Times", 15))
      label.place(x=600,y=400)
      GUI.after(2000,lambda:label.destroy())

    if (arg==4):
      DB.EditMedicinePrice(entry6.get(),entry10.get()) # EditMedicinePrice(name,price)

      label= Label(GUI,text="Price Modified Successfuly",bg="GREY",fg="RED",font=("Times", 15))
      label.place(x=800,y=400)
      GUI.after(2000,lambda:label.destroy())


####################################
############_Receipt_###############
#################################### 
def MakeReceiptUI():
    DestroyAll()
    global label7,entry11,ButtonAddToReceipt, label8,entry12,ButtonMakeReceipt,receiptContents,labelPaymentType,buttoncash,buttonvisa,labelOrderType,buttonstore,buttondelivery
    label7= Label(GUI,text="Enter Medicine Name: ",bg="LightBlue",fg="white",font=("Times", 18),width=20,relief="ridge")
    label7.place(x=360,y=120)
    
    entry11=Entry(GUI , font=("Times", 20),width=20)
    entry11.place(x=350,y=160)
    #######################################
    
    label8= Label(GUI,text="Enter Quantity: ",bg="LightBlue",fg="white",font=("Times", 15),width=15,relief="ridge")
    label8.place(x=670,y=120)
  
    entry12=Entry(GUI , font=("Times", 20),width=15)
    entry12.place(x=650,y=160)
    #######################################
    
    ButtonAddToReceipt = Button(GUI, text ="AddToReceipt",font=("Arial", 14),command = lambda : AddToReceiptUI(entry11.get(),entry12.get()))
    ButtonAddToReceipt.configure(height=1,width=15)
    ButtonAddToReceipt.place(x=550,y=200)
    #######################################
    
    labelPaymentType = Label(GUI,text="Choose Payment Type: ",bg="LightBlue",font=("Arial", 14),relief="ridge")
    labelPaymentType.place(x=400,y=250)
    
    buttoncash=Radiobutton(GUI, text="Cash",variable=PaymentType, value=1,bg="grey",font=("Arial", 14))
    buttoncash.place(x=420,y=280)
    buttonvisa=Radiobutton(GUI,text="Visa",variable=PaymentType, value=2,bg="grey",font=("Arial", 14))
    buttonvisa.place(x=520,y=280)



    labelOrderType = Label(GUI,text="Choose Oder Type: ",bg="LightBlue",font=("Arial", 14),relief="ridge") ## In-Store Or delivery
    labelOrderType.place(x=620,y=250)
    buttonstore=Radiobutton(GUI, text="In-Store",variable=OrderType, value=1,bg="grey",font=("Arial", 14))
    buttonstore.place(x=620,y=280)
    buttondelivery=Radiobutton(GUI,text="Delivery",variable=OrderType, value=2,bg="grey",font=("Arial", 14),command=lambda:DeliveryUI())
    buttondelivery.place(x=720,y=280)
   
    ######################################
    ButtonMakeReceipt = Button(GUI, text ="Generate Receipt",font=("Arial", 20),command = lambda : MakeReceipt())
    ButtonMakeReceipt.configure(height=1,width=20)
    ButtonMakeReceipt.place(x=550,y=500)
    #####################################
    receiptContents = Text(GUI, height=40, width=30)
    receiptContents.insert(END,"Item           "+"Q   "+"Price\n" )


def AddToReceiptUI(MedName,Quantity): # make initial look of receipt contents
  MedDB = MedicineDatabase()
  PriceFromDB = MedDB.SearchMedicineByName(MedName)["price"]
  QuantityInDB= int(MedDB.SearchMedicineByName(MedName)["quantity"])
  
  receiptContents.place(x=900,y=120)
  
  if (int(Quantity)>QuantityInDB):
    ShowError("Insufficient Quantity\n only "+ str(QuantityInDB) +" in stock")
  if (PriceFromDB!="Not Found" and int(Quantity)<=QuantityInDB ):
    receiptContents.insert(END, MedName+space(MedName,15)+Quantity+space(Quantity,5)+PriceFromDB+"\n")

def DeliveryUI():
    global labelClientID,entryClientID,labelAddress,EntryAddress
    labelClientID = Label(GUI,text="Enter Client-ID: ",bg="LightBlue",font=("Arial", 14),relief="ridge")
    labelClientID.place(x=620,y=320)
    entryClientID=Entry(GUI , font=("Times", 20),width=9)
    entryClientID.place(x=620,y=350)

    labelAddress = Label(GUI,text="or Enter Address: ",bg="LightBlue",font=("Arial", 14),relief="ridge")
    labelAddress.place(x=620,y=400)
    EntryAddress=Entry(GUI , font=("Times", 20),width=12)
    EntryAddress.place(x=620,y=430)


def MakeReceipt(): #construct receipt object
  OrderList=receiptContents.get("2.0",END).split("\n")[:-2] #Make list of lines starting from line2 > item:quantiy:price
  receipt = Receipt()
  receipt.items.clear()      #######
  receipt.quantities.clear() #######
  receipt.prices.clear()    #######
  
  if (PaymentType.get()==1):  ###### cash or visa ??
    receipt.SetPaymentType("Cash")
  elif(PaymentType.get()==2):
     receipt.SetPaymentType("Visa")

     

  if (OrderType.get()==1): ## in store or delivery ?
    receipt.SetType("In-Store")
  elif(OrderType.get()==2):
    if(EntryAddress.get()!=" "):
      receipt.SetDeliveryAddress( EntryAddress.get() ) # set address entered manually
    else:
      receipt.SetType("Delivery")
      clientdb=ClientDatabase()
      receipt.SetDeliveryAddress( clientdb.GetClientAddress(entryClientID.get()) )  # setting delivery address with client id given
      receipt.SetClientID(entryClientID.get()) 

 
     

  for i in range(0,len(OrderList)): # add items to receipt object
    item=OrderList[i].split()[0]
    quantity=OrderList[i].split()[1]
    price=OrderList[i].split()[2]
    receipt.AddItem(item,quantity,price)

  GenerateReceipt(receipt,len(OrderList))
  receiptContents.destroy()

   
def GenerateReceipt(receipt,receiptlength): # make file for receipt and preview it
  
 with open("receipt.txt", "w") as receiptfile:
    receiptfile.write("    Group25 Pharmacy\n")
    receiptfile.write("========================\n")
    receiptfile.write("       receipt#"+str(receipt.receiptnum)+"\n")
    receiptfile.write("Payment: "+str(receipt.PaymentType)+"\n")

    if (receipt.DeliveryAddress!=""):
      receiptfile.write("Address: "+str(receipt.DeliveryAddress)+"\n")

    
    receiptfile.write("========================\n")



    receiptfile.write("Item           Q   Price\n")
    receiptfile.write("------------------------\n")
    for i in range(0,receiptlength): ## write all items in receipt
      item=receipt.items[i]
      quantity=receipt.quantities[i]
      price=receipt.prices[i]
      receiptfile.write(item+space(item,15)+quantity+space(quantity,5)+price+"\n")
      
    receiptfile.write("========================\n")
      
    totalprice=str(receipt.CalcSum())    
    receiptfile.write("Total:"+space(totalprice,12)+totalprice+" L.E\n")

    receiptfile.write("========================\n")
    
    receiptfile.write("Thank You For Your Visit!\n")
    receiptfile.write("   "+str(datetime.datetime.now())[:16]+"\n")



 with open("receipt.txt", "r") as receiptfile:
    ReceiptWindow = Tk()
    ReceiptWindow.minsize(300,600)
    Label(ReceiptWindow, text=receiptfile.read()).pack()

    ButtonPrint=Button(ReceiptWindow, text ="Print",font=("Arial", 13),command = lambda : os.startfile("receipt.txt", "print"))
    ButtonPrint.configure(height=1,width=10)
    ButtonPrint.place(x=30,y=500)
    
    ButtonClose=Button(ReceiptWindow, text ="Close",font=("Arial", 13),command = lambda : ReceiptWindow.destroy())
    ButtonClose.configure(height=1,width=10)
    ButtonClose.place(x=160,y=500)     

 receipt.AddToDataBase()

####################################
####################################
####################################


####################################
############_Profit_################
####################################
def DailyProfitUI():
    ooDB = OrderDatabase()
    profit = ooDB.DailyProfit()

    text1.place(x=600,y=200)
    text1.insert(END,str(profit)+"LE")
def MonthlyProfitUI():
    ooDB = OrderDatabase()
    profit = ooDB.MonthlyProfit()

    text2.place(x=800,y=200)
    text2.insert(END, str(profit)+"LE")
def ProfitButtons():
    global Button1, Button2,text2,text1
    DestroyAll()
    Button1 = Button(GUI, text ="Daily profit",font=("Arial", 12),command=lambda: DailyProfitUI())
    Button1.configure(height=2, width=16)
    Button1.place(x=570, y=150)
    Button2 = Button(GUI, text="Monthly profit", font=("Arial", 12), command=lambda: MonthlyProfitUI())
    Button2.configure(height=2, width=16)
    Button2.place(x=770,y=150)
    text1 = Text(master=GUI, height=1, width=10,font=("Arial",12))
    text2 = Text(master=GUI, height=1, width=10,font=("Arial",12))


####################################
####################################
####################################

def AddClientUI():
    global label9,label10,labe11,entry99,entry100,entry111,ButtonAddClient
    DestroyAll()

    label9= Label(GUI,text="Name",    bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    label9.place(x=700,y=120)
    label10= Label(GUI,text="Address", bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    label10.place(x=700,y=160)
    labe11= Label(GUI,text="Phone",bg="LightBlue",fg="white",font=("Times", 20),width=7,relief="ridge")
    labe11.place(x=700,y=200)

    
    entry99=Entry(GUI , font=("Times", 20))
    entry99.place(x=850,y=120)
    entry100=Entry(GUI , font=("Times", 20))
    entry100.place(x=850,y=160)
    entry111=Entry(GUI , font=("Times", 20))
    entry111.place(x=850,y=200)

    ButtonAddClient = Button(GUI, text ="Add",font=("Arial", 14),command = lambda : NewClient())
    ButtonAddClient.configure(height=1,width=10)
    ButtonAddClient.place(x=800,y=250)

def NewClient():    
  client=Client()
  client.SetName(entry99.get())
  client.SetAddress(entry100.get())
  client.SetPhone(entry111.get())   
  client.AddToDataBase()
  labeldone= Label(GUI,text="Client Added Successfuly, With ID: "+str(client.ID),bg="GREY",fg="RED",font=("Times", 20)) 
  labeldone.place(x=700,y=360)
  GUI.after(4000,lambda:labeldone.destroy()) # done label appear and disappear after time

    
####################################
####################################
####################################

def ReturnOrderUI():
    DestroyAll()
    global label1,entry2,Button1
    label1= Label(GUI,text="Enter Order ID: ",bg="LightBlue",fg="white",font=("Times", 18),width=15,relief="ridge")
    label1.place(x=760,y=120)
    
    entry2=Entry(GUI , font=("Times", 20),width=13)
    entry2.place(x=770,y=160)

        
    Button1 = Button(GUI, text ="Return",font=("Arial", 14),command = lambda : ReturnOrder())
    Button1.configure(height=1,width=14)
    Button1.place(x=780,y=200)

def ReturnOrder():
  DB=OrderDatabase()
  DB.ReturnOrder(entry2.get())
  
  labeldone= Label(GUI,text="Order Returned Successfully ",bg="GREY",fg="RED",font=("Times", 20)) 
  labeldone.place(x=700,y=360)
  GUI.after(4000,lambda:labeldone.destroy())
  
def main():
 global GUI
 GUI = Tk()
 GUI.title("Pharmacy Managment System")
 GUI.configure(bg='Grey')
 GUI.minsize(1400,650)
 GUI.resizable(0,0)

 try:
   AdminGui.destroy()
 except:
   pass
  
# LoginScreen.destroy()

 global PaymentType
 PaymentType = IntVar() ###############
 global OrderType
 OrderType = IntVar()


 photo=PhotoImage(file = "1.png")
 labelbanner= Label(GUI,text="Pharmacy Management System",bg="LightBlue",fg="white",font=("Times", 30),relief="ridge")
 labelbanner.grid(columnspan=7,padx=500,sticky='ew')

 B0 = Button(GUI, text ="Add New Medicine",font=("Arial", 15),command = lambda : AddMedicineUI())
 B0.configure(height=2,width=16)
 B0.grid(row=1,column=0)

 B1 = Button(GUI, text ="Edit Existing Medicine",font=("Arial", 15), command =lambda :  EditExistingMedicineUI())
 B1.configure(height=2,width=17)
 B1.grid(row=1,column=1)

 B2 = Button(GUI, text ="Make A Receipt",font=("Arial", 15), command =lambda :  MakeReceiptUI())
 B2.configure(height=2,width=16)
 B2.grid(row=1,column=2)

 B3 = Button(GUI, text ="Income",font=("Arial", 15), command=lambda: ProfitButtons())
 B3.configure(height=2,width=16)
 B3.grid(row=1,column=3)

 B4 = Button(GUI, text ="Add Client",font=("Arial", 15), command =lambda :  AddClientUI())
 B4.configure(height=2,width=16)
 B4.grid(row=1,column=4)

 B5 = Button(GUI, text ="Return Order",font=("Arial", 15), command =lambda :  ReturnOrderUI())
 B5.configure(height=2,width=16)
 B5.grid(row=1,column=5)

 B6 = Button(GUI, text ="Administration",font=("Arial", 15), command =lambda :  AdminUi())
 B6.configure(height=2,width=16)
 B6.grid(row=1,column=6)

 labelfooter= Label(GUI,text="Version 1.00",bg="Grey",font=("Times", 14))
 labelfooter.place(x=700,y=600)

 
 GUI.mainloop()
 
 

def login():
  global LoginScreen
  LoginScreen = Tk()
  LoginScreen.title("Pharmacy Managment System")
  LoginScreen.configure(bg='GREY')
  LoginScreen.minsize(1400,650)
  LoginScreen.resizable(0,0)
  
  labelbanner= Label(LoginScreen,text="Login To Continue",bg="RED",fg="BLACK",font=("Times", 30))
  labelbanner.grid(columnspan=7,padx=500)
  
  usernameLabel=Label(LoginScreen, text=" Username: ", width=7, bg="grey",font=("Times",20))
  usernameLabel.place(x=350,y=150)
  usernameEntry = Entry(LoginScreen,font=("Times", 20))
  usernameEntry.place(x=550,y=150)
  
  
  PasswordLabel=Label(LoginScreen, text="Password:" , width=7, bg="grey",font=("Times",20))
  PasswordLabel.place(x=350,y=250)
  passwordEntry = Entry(LoginScreen, font=("Times", 20), show= '*')
  passwordEntry.place(x=550,y=250)

  
  loginbutton= Button(LoginScreen, text ="Login",font=("Arial", 20), command =lambda :  main())
  loginbutton.place(x=600,y=400)
  LoginScreen.mainloop()



####################################
####################################
####################################

class EmployeeDataBase(Database):
    DataBaseFile = Admin_path
    workbook = load_workbook(filename=DataBaseFile)
    sheet = workbook.active

    def SearchEmployee(self, EmployeeName):
        EmployeeDectionary = {}
        for row in range(1, self.sheet.max_row + 1):  # loop over rows
            if (str(self.sheet["C" + str(row)].value).lower() == EmployeeName.lower()):
                EmployeeDectionary["username"] = self.sheet["A" + str(row)].value
                EmployeeDectionary["password"] = self.sheet["B" + str(row)].value
                EmployeeDectionary["national_id"] = self.sheet["D" + str(row)].value
                EmployeeDectionary["address"] = self.sheet["E" + str(row)].value
                EmployeeDectionary["phone"] = self.sheet["F" + str(row)].value
                EmployeeDectionary["age"] = self.sheet["G" + str(row)].value
                return EmployeeDectionary

    def AddEmployee(self,Employee):
      Row = str(self.GetRowIndex())
      if (Employee.name!=""):
        self.sheet["A" + Row] = Employee.username
        self.sheet["B" + Row] = Employee.password
        self.sheet["C" + Row] = Employee.name
        self.sheet["D" + Row] = Employee.national_id
        self.sheet["E" + Row] = Employee.phone_number
        self.sheet["F" + Row] = Employee.address
        self.sheet["G" + Row] = Employee.age
        self.sheet["H" + Row] = Employee.SalaryPerHr
        self.sheet["K" + Row] = 0 #month working hours
        self.sheet["H" + Row] = Employee.SalaryPerHr
        self.sheet["L" + Row] = Employee.privilege
        self.SaveDatabase()

####################################
####################################
####################################

class Employee:
    username = 'null'
    name = 'null'
    national_id = 'null'
    phone_number = 'null'
    address = 'null'
    age = 'null'
    password = 'null'
    SalaryPerHr=0
    privilege='user'

    def SetUserName(self, UserName):
        self.username = UserName.lower()
    def SetPassword(self, Password):
        self.password = Password.lower()
    def SetName(self, Name):
        self.name = Name.lower()
    def SetNationalId(self, Id):
        self.national_id = Id
    def SetPhoneNumber(self, phone):
        self.phone_number = phone
    def SetAddress(self, Add):
        self.address = Add
    def SetAge(self, Age):
        self.age = Age
    def SetSalary(self, salary):
        self.SalaryPerHr = salary
    def SetPrivilege(self, priv):
      self.privilege=priv    
    def AddToDataBase(self):
        AdminDB = EmployeeDataBase()
        AdminDB.AddEmployee(self)
    def CheckIn(self):
      DB=EmployeeDataBase()
      for row in range(1,DB.GetRowIndex()):
        if (str(self.name).lower()==str(DB.sheet["C"+str(row)].value).lower()):
          DB.sheet["I"+str(row)]=str(datetime.datetime.now())[:16]
          DB.SaveDatabase()
    def CheckOut(self):
      DB=EmployeeDataBase()
      for row in range(1,DB.GetRowIndex()):
        if (str(self.name).lower()==str(DB.sheet["C"+str(row)].value).lower()):
          DB.sheet["J"+str(row)]=str(datetime.datetime.now())[:16]
          DB.SaveDatabase()
          checkin=DB.sheet["I"+str(row)].value
          checkout=DB.sheet["J"+str(row)].value
        #  workedtoday=datetime.datetime.strptime(checkout,'%Y-%m-%d %H:%M')-datetime.datetime.strptime(checkin,'%Y-%m-%d %H:%M')
          workedtoday=int(checkout[11:13])-int(checkin[11:13])
          monthwork=DB.sheet["K"+str(row)].value #monthly working hours
          DB.sheet["K"+str(row)]=monthwork+workedtoday
          DB.SaveDatabase()
    def CalcMonthSalary(self):
      DB=EmployeeDataBase()
      for row in range(1,DB.GetRowIndex()):
        if (str(self.name).lower()==str(DB.sheet["C"+str(row)].value).lower()):
           monthwork=int(DB.sheet["K"+str(row)].value)
           monthsalary=int(DB.sheet["H"+str(row)].value)
           return (monthsalary*monthwork)
    def GetWorkedHrs(self):
      DB=EmployeeDataBase()
      for row in range(1,DB.GetRowIndex()):
           if (str(self.name).lower()==str(DB.sheet["C"+str(row)].value).lower()):
               monthwork=DB.sheet["K"+str(row)].value
               return (monthwork)
   

def NewEmployee():
    employee = Employee()
    employee.SetUserName(entry6.get())
    employee.SetPassword(entry7.get())
    employee.SetName(entry8.get())
    employee.SetNationalId(entry9.get())
    employee.SetAddress(entry10.get())
    employee.SetPhoneNumber(entry11.get())
    employee.SetAge(entry12.get())
    employee.SetSalary(entry13.get())
    employee.SetPrivilege(Priv.get())
    employee.AddToDataBase()
    employeeadded = Label(AdminGui, text="Employee Added Successfuly", bg="GREY", fg="Green", font=("Times", 20))
    employeeadded.place(x=120, y=480)
    AdminGui.after(2000,lambda:employeeadded.destroy())

############################
############################
############################

def DisplayEmployeeInfo(searched_name):
    global info_display

    D2 = EmployeeDataBase()
    Dictdata = D2.SearchEmployee(searched_name)
    
    info_display = Text(AdminGui, height=15, width=25)
    info_display.grid(column=2, row=5)
    info_display.insert(INSERT, 'Name: '+searched_name+'\n')
    info_display.insert(INSERT, 'User Name: ' + Dictdata["username"] + '\n')
    info_display.insert(INSERT, 'Password: ' + Dictdata["password"] + '\n')
    info_display.insert(INSERT, 'National ID: ' + Dictdata["national_id"] + '\n')
    info_display.insert(INSERT, 'Address: ' + Dictdata["address"] + '\n')
    info_display.insert(INSERT, 'Phone: ' + Dictdata["phone"] + '\n')
    info_display.insert(INSERT, 'Age: ' + Dictdata["age"] + '\n')
    info_display.configure(state='disabled')


def GetEmployeeName():
    global search_label,search_entry,Buttonseach
    DestroyAll()

    search_label = Label(AdminGui, text="Name", bg="Green", fg="white", font=("Times", 20), width=9, relief="ridge")
    search_label.place(x=0, y=120)
    search_entry = Entry(AdminGui, font=("Times", 20))
    search_entry.place(x=150, y=120)
    Buttonseach = Button(AdminGui, text="Search", font=("Arial", 14), command=lambda: DisplayEmployeeInfo(search_entry.get()))
    Buttonseach.configure(height=1, width=10)
    Buttonseach.place(x=100, y=160)

############################
############################
############################
    
def AddEmployee():
    global label6, label7, label8, label9, labe20,labe21,labe22, entry6, entry7, entry8, entry9, entry10,entry11,entry12, AddInfo,label33,entry13,buttonAdmin,buttonuser
    global buttonAdmin,buttonuser
    DestroyAll()

    label6 = Label(AdminGui, text="Username", bg="Green", fg="white", font=("Times", 20), width=9, relief="ridge")
    label6.place(x=0, y=120)
    label7 = Label(AdminGui, text="Password", bg="Green", fg="white", font=("Times", 20), width=9, relief="ridge")
    label7.place(x=0, y=160)
    label8 = Label(AdminGui, text="Name", bg="Green", fg="white", font=("Times", 20), width=9, relief="ridge")
    label8.place(x=0, y=200)
    label9 = Label(AdminGui, text="National ID", bg="Green", fg="white", font=("Times", 20), width=9, relief="ridge")
    label9.place(x=0, y=240)
    labe20 = Label(AdminGui, text="Address", bg="Green", fg="white", font=("Times", 20), width=9, relief="ridge")
    labe20.place(x=0, y=280)
    labe21 = Label(AdminGui, text="Phone Num", bg="Green", fg="white", font=("Times", 20), width=9, relief="ridge")
    labe21.place(x=0, y=320)
    labe22 = Label(AdminGui, text="Age", bg="Green", fg="white", font=("Times", 20), width=9, relief="ridge")
    labe22.place(x=0, y=360)
    label33 = Label(AdminGui, text="Salary/Hr", bg="Green", fg="white", font=("Times", 20), width=9, relief="ridge")
    label33.place(x=0, y=400)
    
    entry6 = Entry(AdminGui, font=("Times", 20))
    entry6.place(x=150, y=120)
    entry7 = Entry(AdminGui, font=("Times", 20))
    entry7.place(x=150, y=160)
    entry8 = Entry(AdminGui, font=("Times", 20))
    entry8.place(x=150, y=200)
    entry9 = Entry(AdminGui, font=("Times", 20))
    entry9.place(x=150, y=240)
    entry10 = Entry(AdminGui, font=("Times", 20))
    entry10.place(x=150, y=280)
    entry11 = Entry(AdminGui, font=("Times", 20))
    entry11.place(x=150, y=320)
    entry12 = Entry(AdminGui, font=("Times", 20))
    entry12.place(x=150, y=360)
    entry13 = Entry(AdminGui, font=("Times", 20))
    entry13.place(x=150, y=400)     

    
    buttonuser=Radiobutton(AdminGui,text="User", variable=Priv ,  value='user',  bg="grey",font=("Arial", 14))
    buttonuser.place(x=250,y=440)
    
    buttonAdmin=Radiobutton(AdminGui,text="Administrator",variable=Priv, value='admin',bg="grey",font=("Arial", 14))
    buttonAdmin.place(x=350,y=440)
    
    
    AddInfo = Button(AdminGui, text="Add", font=("Arial", 14), command=lambda: NewEmployee())
    AddInfo.configure(height=1, width=10)
    AddInfo.place(x=100, y=440)
    
############################
############################
############################
def CheckInUI():
    global EnterName,ListBox,ButtonCheckin,ButtonCheckOut
    DestroyAll()

    EnterName = Label(AdminGui, text="Select Employee", bg="Green", fg="white", font=("Times", 15), width=18, relief="ridge")
    EnterName.place(x=400, y=120)
	
    DB=EmployeeDataBase()
    ListBox = Listbox(AdminGui)
    for row in range(2,DB.GetRowIndex()):
      ListBox.insert(row,DB.sheet["C"+str(row)].value)
    ListBox.configure(font=("Times", 15))  
    ListBox.place(x=400,y=160)
    
#print (ListBox.get(ListBox.curselection()))
    
    ButtonCheckin = Button(AdminGui, text="Check-In", font=("Arial", 14), command=lambda: CheckIn() )
    ButtonCheckin.configure(height=1, width=8)
    ButtonCheckin.place(x=650, y=160)
	
    ButtonCheckOut = Button(AdminGui, text="Check-Out", font=("Arial", 14), command=lambda: CheckOut())
    ButtonCheckOut.configure(height=1, width=8)
    ButtonCheckOut.place(x=650, y=240)	
	
def CheckIn():
 #try:
  employee=Employee()
  employee.SetName(ListBox.get(ListBox.curselection()))
  employee.CheckIn()
  labeldone = Label(AdminGui, text="Employee Check-In Recoreded Successfully", bg="GREY", fg="GREEN", font=("Times", 15))
  labeldone.place(x=650, y=200)
  AdminGui.after(2000,lambda:labeldone.destroy())
# except:
 #  ShowError("Please Select Employee")
   
def CheckOut():
# try: 
  employee=Employee()
  employee.SetName(ListBox.get(ListBox.curselection()))
  employee.CheckOut()
  
  labeldone = Label(AdminGui, text="Employee Check-Out Recoreded Successfully", bg="GREY", fg="GREEN", font=("Times", 15))
  labeldone.place(x=650, y=280)
  AdminGui.after(2000,lambda:labeldone.destroy())
 #except:
 #  ShowError("Please Select Employee")



def WorkingHoursUI():
  global EnterName,ListBox,Button1,Button2
  DestroyAll()

  EnterName = Label(AdminGui, text="Select Employee", bg="Green", fg="white", font=("Times", 15), width=18, relief="ridge")
  EnterName.place(x=400, y=120)
	
  DB=EmployeeDataBase()
  ListBox = Listbox(AdminGui)
  for row in range(2,DB.GetRowIndex()):
      ListBox.insert(row,DB.sheet["C"+str(row)].value)
  ListBox.configure(font=("Times", 15))  
  ListBox.place(x=400,y=160)

  Button1 = Button(AdminGui, text="Current-Month Working Hours", font=("Arial", 14), command=lambda: WorkingHrs() )
  Button1.configure(height=1, width=25)
  Button1.place(x=650, y=160)
	
  Button2 = Button(AdminGui, text="Deserved Salary", font=("Arial", 14), command=lambda: CalcSalary())
  Button2.configure(height=1, width=14)
  Button2.place(x=650, y=240)  

def WorkingHrs():
  employee=Employee()
  employee.SetName(ListBox.get(ListBox.curselection()))
  hrs = employee.GetWorkedHrs()

  labelhours = Label(AdminGui, text="Employee Worked "+str(hrs)+" Hrs This Month",bg="GREY", fg="Blue", font=("Times", 15))
  labelhours.place(x=650, y=200)
  AdminGui.after(5000,lambda:labelhours.destroy())

def CalcSalary():
  employee=Employee()
  employee.SetName(ListBox.get(ListBox.curselection())) 
  salary = employee.CalcMonthSalary()
  labelsalary = Label(AdminGui, text="Employee Deserves "+str(salary)+" L.E This Month",bg="GREY", fg="Blue", font=("Times", 15))
  labelsalary.place(x=650, y=280)
  AdminGui.after(5000,lambda:labelsalary.destroy())
  
############################
############################
############################
   
def AdminUi():
    GUI.destroy()
    
    global AdminGui
    AdminGui = Tk()
    
    global Priv
    Priv = StringVar()

    AdminGui.title("Adminstration mode")
    AdminGui.configure(bg='GREY')
    AdminGui.minsize(1400, 650)
    AdminGui.resizable(0, 0)
    banner = Label(AdminGui, text="Adminstration mode", bg="Green", fg="white", font=("Times", 30),relief="ridge")
    banner.grid(columnspan=7, padx=500)

    
 

    Bot1 = Button(AdminGui, text="Add New Employee", font=("Arial", 15), command=lambda: AddEmployee())
    Bot1.configure(height=2, width=16)
    Bot1.grid(row=1, column=0)

    Bot2 = Button(AdminGui, text="Employee Information", font=("Arial", 14), command=lambda: GetEmployeeName())
    Bot2.configure(height=2, width=17)
    Bot2.grid(row=1, column=1)
	
    Bot3 = Button(AdminGui, text="Check-in/out", font=("Arial", 15), command=lambda: CheckInUI())
    Bot3.configure(height=2, width=16)
    Bot3.grid(row=1, column=2)


    Bot4 = Button(AdminGui, text="Working Hours", font=("Arial", 15), command=lambda: WorkingHoursUI())
    Bot4.configure(height=2, width=16)
    Bot4.grid(row=1, column=3)

    Bot5 = Button(AdminGui, text="Sales", font=("Arial", 15), command=lambda: WorkingHoursUI())
    Bot5.configure(height=2, width=16)
    Bot5.grid(row=1, column=4)

    Bot6 = Button(AdminGui, text="Back", font=("Arial", 15), command=lambda:  main())
    Bot6.configure(height=2, width=16)
    Bot6.grid(row=1, column=5)

    AdminGui.mainloop()




def DestroyAll(): # make sure that area we use is clear before placing objects
  try:
        label1.destroy()#add medicine
        label2.destroy()
        label3.destroy()
        label4.destroy()
        label5.destroy()
        entry1.destroy()
        entry2.destroy()
        entry3.destroy()
        entry4.destroy()
        entry5.destroy()
        ButtonAdd.destroy()
  except:
           pass
  try:
        label6.destroy() #edit Medicine
        entry6.destroy()
        ButtonSearch.destroy()
        entry7.destroy()
        entry8.destroy()
        entry9.destroy()
        entry10.destroy()
        ButtonBarcode.destroy()
        ButtonQuantity.destroy()
        ButtonExpire.destroy()
        ButtonPrice.destroy()
  except:
            pass
  try:  
        label7.destroy() #receipt
        entry11.destroy()
        ButtonAddToReceipt.destroy()
        label8.destroy()
        entry12.destroy()
        ButtonMakeReceipt.destroy()
        buttoncash.destroy()
        buttonvisa.destroy()
        labelPaymentType.destroy()
        receiptContents.destroy()
        labelOrderType.destroy()
        buttonstore.destroy()
        buttondelivery.destroy()
        labelClientID.destroy()
        entryClientID.destroy()
        labelAddress.destroy()
        EntryAddress.destroy()

  except:
         pass
  try:
      Button1.destroy() ##add client
      Button2.destroy()
      text1.destroy()
      text2.destroy()
  except:
        pass
  try:
    label1.destroy() #return order
    entry2.destroy()
    Button1.destroy()
  except:
    pass
  try:  #Add Client
    label9.destroy()
    label10.destroy()
    labe11.destroy()
    entry99.destroy()
    entry100.destroy()
    entry111.destroy()
    ButtonAddClient.destroy()
  except:
    pass
  try: #add employee
        label6.destroy()
        label7.destroy()
        label8.destroy()
        label9.destroy()
        labe20.destroy()
        labe21.destroy()
        labe22.destroy()
        label33.destroy()
        entry6.destroy()
        entry7.destroy()
        entry8.destroy()
        entry9.destroy()
        entry10.destroy()
        entry11.destroy()
        entry12.destroy()
        entry13.destroy()
        AddInfo.destroy()
  except:
        pass
  try:
    search_label.destroy()
    search_entry.destroy()
    Buttonseach.destroy()
    info_display.destroy()
  except:
    pass
  try: #checkin
     EnterName.destroy()
     ListBox.destroy()
     ButtonCheckin.destroy()
     ButtonCheckOut.destroy()
  except:
    pass
class FullScreenApp(object):
    def __init__(self, master, **kwargs):
        self.master=master
        pad=3
        self._geom='200x200+0+0'
        master.geometry("{0}x{1}+0+0".format(
            master.winfo_screenwidth()-pad, master.winfo_screenheight()-pad))
        master.bind('<Escape>',self.toggle_geom)            
    def toggle_geom(self,event):
        geom=self.master.winfo_geometry()
        print(geom,self._geom)
        self.master.geometry(self._geom)
        self._geom=geom

main()



'''receipt=Receipt()
receipt.AddItem("mon",12,14)
receipt.AddItem("ahmed",10,20)
receipt.AddItem("_",3,2)
receipt.printrec()
receipt.CalcSum()

for key,value in Receipt.items.items():
    print (key[0])'''
        

